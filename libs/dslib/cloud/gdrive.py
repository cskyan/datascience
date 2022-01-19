#!/usr/bin/env python
# -*- coding=utf-8 -*-
###########################################################################
# Copyright (C) 2013-2021 by Caspar. All rights reserved.
# File Name: gdrive.py
# Author: Shankai Yan
# E-mail: dr.skyan@gmail.com
# Created Time: 2021-12-10 23:36:38
###########################################################################
#

import os, json
from google.colab import auth
from google.colab import drive
from oauth2client.client import GoogleCredentials

from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

import gspread
import pandas as pd
from openpyxl.utils.cell import get_column_letter


GDRIVE_ROOT_PATH = '/content/gdrive'


def mkdir(path):
  if path and not os.path.exists(path):
    print(("Creating folder: " + path))
    os.makedirs(path)
    drive.flush_and_unmount() # To be replace by drive.flush()
    drive.mount(GDRIVE_ROOT_PATH)

mkdir(GDRIVE_ROOT_PATH)
auth.authenticate_user()
drive.mount(GDRIVE_ROOT_PATH)
GDRIVE_PATH = os.path.join(GDRIVE_ROOT_PATH, 'MyDrive')

DATA_ROOT_PATH = os.path.join(GDRIVE_PATH, 'notebooks/data')
DATA_ROOT_ID = ''

gcred = GoogleCredentials.get_application_default()
gauth = GoogleAuth()
gauth.credentials = gcred
GDRIVE = GoogleDrive(gauth)
GDRIVE_FOLDER_TYPE = 'application/vnd.google-apps.folder'
GDRIVE_SPREADSHEET_TYPE = 'application/vnd.google-apps.spreadsheet'

GSC = gspread.authorize(gcred)

GMETA_TEMPLATE = {'spreadsheet':{}}


class GMeta(object):
  def __init__(self, meta_fpath):
    self.meta_fpath = meta_fpath
    if not os.path.exists(meta_fpath):
      self.meta_dict = GMETA_TEMPLATE
      self.flush()
    else:
      with open(meta_fpath, 'r') as fd:
        self.meta_dict = json.load(fd)
  def get_spreadsheet(self, fpath):
    dir_path, gs_name = os.path.split(fpath)
    gs_name = os.path.splitext(gs_name)[0]
    fpath = os.path.join(dir_path, '%s.gsheet' % gs_name)
    return self.meta_dict['spreadsheet'][fpath] if fpath in self.meta_dict['spreadsheet'] else None
  def add_spreadsheet(self, fpath, gsheet_id):
    self.meta_dict['spreadsheet'][fpath] = gsheet_id
    self.flush()
  def flush(self, **kwargs):
    kw_args = dict(sort_keys=True, indent=4, separators=(',', ': '))
    kw_args.update(kwargs)
    with open(self.meta_fpath, mode='w') as fd:
      fd.write(json.dumps(self.meta_dict, **kw_args))


def rect2range(rect):
  return '%s%i:%s%i' % (get_column_letter(rect[1]+1), rect[0]+1, get_column_letter(rect[1]+rect[3]), rect[0]+rect[2])

def get_gsheet(gsheet, sheet='sheet1'):
  if type(sheet) is int:
    worksheet = gsheet.get_worksheet(sheet)
  elif type(sheet) is str and sheet != 'sheet1':
    try:
      worksheet = gsheet.worksheet(sheet)
    except Exception as e:
      worksheet = gsheet.sheet1
  else:
    worksheet = gsheet.sheet1
  return worksheet

def to_gsheet(dataframe, fpath=None, folder_id=None, sheet='sheet1', columns=None, header=True, index=True, index_label=None, gmeta=None):
  if fpath is None:
    fpath = os.path.join(DATA_PATH, 'new_spreadsheet')
  dir_path, gs_name = os.path.split(fpath)
  # Fix directory prefix
  if not dir_path.startswith(DATA_ROOT_PATH):
    fpath_nodes = dir_path.strip('/').split('/')
    default_nodes = set(DATA_ROOT_PATH.strip('/').split('/'))
    dir_path = '/'.join([DATA_ROOT_PATH] + [n for n in fpath_nodes[:len(default_nodes)] if n not in default_nodes] + fpath_nodes[len(default_nodes):])
  gs_name = os.path.splitext(gs_name)[0]
  fpath = os.path.join(dir_path, '%s.gsheet' % gs_name)
  mkdir(dir_path)
  found_folder_id = folder_id
  # Find the folder ID
  if folder_id is None:
    rel_nodes = dir_path[len(DATA_ROOT_PATH):].strip('/').split('/')
    search_folder_id = DATA_ROOT_ID
    while len(rel_nodes) > 0:
      try:
        search_folder_id = next(f for f in GDRIVE.ListFile({'q': "'%s' in parents" % search_folder_id}).GetList() if f['mimeType'] == GDRIVE_FOLDER_TYPE and f['title'] == rel_nodes[0])['id']
      except Exception as e:
        print('Path [%s] not found!' % dir_path)
        break
      del rel_nodes[0]
    else:
      found_folder_id = search_folder_id
  # Get worksheet
  gsheet = GSC.create(gs_name, folder_id=found_folder_id)
  worksheet = get_gsheet(gsheet, sheet=sheet)

  # Determine the range of sheet columns, index, data and update the values
  if header:
    columns = dataframe.columns if columns is None else columns
  if index:
    index_label = dataframe.index.name if index_label is None else index_label
  if header and index:
    worksheet.update('A1', index_label)
    worksheet.update(rect2range((0, 1, 1, len(columns))), [columns.tolist()])
    worksheet.update(rect2range((1, 0, len(dataframe.index), 1)), [[idx] for idx in dataframe.index])
    rect = (1, 1) + dataframe.shape
  elif header:
    worksheet.update(rect2range((0, 0, 1, len(columns))), [columns.tolist()])
    rect = (1, 0) + dataframe.shape
  elif index:
    worksheet.update('A1', index_label)
    worksheet.update(rect2range((1, 0, len(dataframe.index), 1)), [[idx] for idx in dataframe.index])
    rect = (0, 1) + dataframe.shape
  else:
    rect = (0, 0) + dataframe.shape
  worksheet.update(rect2range(rect), dataframe.values.tolist())
  print('Saved to %s' % ((GDRIVE_PATH if found_folder_id is None else 'fpath: [%s]' % fpath) if folder_id is None else 'Folder ID: %s' % folder_id))
  if folder_id is None and gmeta is not None:
    try:
      gmeta.add_spreadsheet(fpath, worksheet.spreadsheet.id)
    except Exception as e:
      print(e)
  return worksheet.spreadsheet.id


def read_gsheet(fpath=None, gsheet_id=None, sheet='sheet1', header=True, names=None, index_col=None, usecols=None, dtype=None, gmeta=None):
  worksheet = None
  # Get worksheet
  if fpath is None and gsheet_id is None:
    print('Please input a correct file path or sheet ID!')
    return None
  elif gsheet_id is not None:
    try:
      gsheet = GSC.open_by_key(gsheet_id)
      worksheet = get_gsheet(gsheet, sheet=sheet)
    except Exception as e:
      if fpath is None:
        print('Inexisted sheet ID: %s' % gsheet_id)
        return None
  if worksheet is None:
    dir_path, gs_name = os.path.split(fpath)
    # Fix directory prefix
    if not dir_path.startswith(DATA_ROOT_PATH):
      print('Path [%s] is not within the data root path [%s]. Trying to load from the data root path. Next time consider inputting the sheet ID instead!' % (fpath, DATA_ROOT_PATH))
      fpath_nodes = dir_path.strip('/').split('/')
      default_nodes = set(DATA_ROOT_PATH.strip('/').split('/'))
      dir_path = '/'.join([DATA_ROOT_PATH] + [n for n in fpath_nodes[:len(default_nodes)] if n not in default_nodes] + fpath_nodes[len(default_nodes):])
    gs_name = os.path.splitext(gs_name)[0]
    fpath = os.path.join(dir_path, '%s.gsheet' % gs_name)
    if not os.path.exists(fpath):
      print('File path [%s] cannot be found!' % fpath)
      return None
    rel_nodes = dir_path[len(DATA_ROOT_PATH):].strip('/').split('/')
    search_folder_id, found_folder_id = DATA_ROOT_ID, None
    while len(rel_nodes) > 0:
      try:
        search_folder_id = next(f for f in GDRIVE.ListFile({'q': "'%s' in parents" % search_folder_id}).GetList() if f['mimeType'] == GDRIVE_FOLDER_TYPE and f['title'] == rel_nodes[0])['id']
      except Exception as e:
        print('Path [%s] not found!' % dir_path)
        return None
      del rel_nodes[0]
    else:
      found_folder_id = search_folder_id
    try:
        gsheet_id = next(f for f in GDRIVE.ListFile({'q': "'%s' in parents" % found_folder_id}).GetList() if f['mimeType'] == GDRIVE_SPREADSHEET_TYPE and f['title'] == gs_name)['id']
    except Exception as e:
      print('Sheet [%s] is not found in the path [%s]!' % (gs_name, dir_path))
      return None
    if gmeta is not None:
      try:
        gmeta.add_spreadsheet(fpath, gsheet_id)
      except Exception as e:
        print(e)
    gsheet = GSC.open_by_key(gsheet_id)
    worksheet = get_gsheet(gsheet, sheet=sheet)

  sheet_values = worksheet.get_all_values()
  if header:
    if sheet_values[0][0] == '':
      sheet_values[0][0] = '_index'
      index_col = index_col if index_col else '_index'
    columns = sheet_values[0] if names is None else names
    dataframe = pd.DataFrame.from_records(sheet_values[1:])
    dataframe.columns = columns
  else:
    dataframe = pd.DataFrame.from_records(sheet_values)
  if index_col:
    dataframe = dataframe.set_index(index_col)
  if usecols:
    dataframe = dataframe[usecols]
  if dtype:
    for k, v in dtype.items():
      try:
        dataframe[k] = dataframe[k].astype(v)
      except Exception as e:
        print('Cannot set the dtype of column %s as %s!' % (k, v))
  return dataframe


def batch_pd2gsh(dir_path, read_kwargs={}, write_kwargs={}):
    if not os.path.exists(dir_path): return []
    return [to_gsheet(pd.read_csv(os.path.join(dir_path, fname), **read_kwargs), os.splitext(fname)[0], **write_kwargs) for fname in fs.listf(dir_path, pattern='.*\.csv', ret_fullpath=False)]


def batch_gsh2pd(dir_path=None, gsheet_ids=None, read_kwargs={}, write_kwargs={}):
    return [read_gsheet(os.path.join(dir_path, fname), **read_gsheet).to_csv('%s.csv'%os.splitext(fname)[0]) for fname in fs.listf(dir_path, pattern='.*\.gsheet', ret_fullpath=False)]


def main():
    pass


if __name__ == '__main__':
    main()
